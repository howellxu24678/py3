# -*- coding: utf-8 -*-
__author__ = 'xujh'

#打包 pyinstaller -F --upx-dir D:\soft\upx391w -i excel_72pxt.ico auto_excel.py
#pip install PyQt5 http://pypi.douban.com/simple --trusted-host=pypi.douban.com
import os
import logging.config
import configparser
from openpyxl import *
import re
from openpyxl.styles import *
import copy
import time

def wait_to_quit():
    print("按回车键结束")
    if input():
        exit()


class AutoExcel(object):
    def __init__(self, configfile):
        try:
            self._cf = configparser.ConfigParser()
            self._cf.read(os.path.join(os.getcwd(), baseconfdir, configfile), encoding='UTF-8')

            self._txt_file_path = self._cf.get("path", "txt_file_path")
            if not os.path.exists(self._txt_file_path):
                logger.error("txt文件不存在，文件路径配置为:%s", self._txt_file_path)
                return wait_to_quit()

            self._xlsx_file_path = self._cf.get("path", "xlsx_file_path")
            if not os.path.exists(self._xlsx_file_path):
                logger.info("xlsx文件不存在，将创建一个新的，文件路径配置为：%s", self._xlsx_file_path)
                dirname = os.path.dirname(self._xlsx_file_path)
                if not os.path.exists(dirname):
                    os.mkdir(dirname)
                Workbook().save(self._xlsx_file_path)

            # 每个类型的记录对应提取的结果集个数（用于检查记录是否正确）
            self._dict_title_values_count = {}
            for vc in self._cf.options("values_count"):
                self._dict_title_values_count[vc] = self._cf.getint("values_count", vc)

            # 结果集为数值的位置
            self._dict_title_float_pos = {}
            for fp in self._cf.options("float_pos"):
                self._dict_title_float_pos[fp] = [int(x) for x in self._cf.get("float_pos", fp).split(',')]

            # 结果集校验
            self._dict_check = {}
            for ch in self._cf.options("check"):
                self._dict_check[ch] = self._cf.get("check", ch).strip()

            # 记录类型及需要追加记录的excel表格名
            self._dict_title_add_sheet = {}
            # 记录类型及需要写入的列数据
            self._dict_add_sheet_column = {}
            # 记录类型及需要更新记录的excel表格名
            self._dict_title_update_sheet = {}
            # 记录类型及需要更新的列
            self._dict_update_sheet_column = {}
            # 更新的记录需要满足的条件
            self._dict_update_sheet_condition = {}


            for asn in self._cf.options("add_sheet"):
                self._dict_title_add_sheet[asn] = self._cf.get("add_sheet", asn).strip()
                d = {}
                for asc in self._cf.options("add_sheet:" + asn):
                    d[int(asc)] = self._cf.get("add_sheet:" + asn, asc).strip()
                self._dict_add_sheet_column[asn] = d

            for usn in self._cf.options("update_sheet"):
                self._dict_title_update_sheet[usn] = self._cf.get("update_sheet", usn).strip()
                d = {}
                for usc in self._cf.options("update_sheet:" + usn):
                    d[int(usc)] = self._cf.get("update_sheet:" + usn, usc).strip()
                self._dict_update_sheet_column[usn] = d

            d_match = {}
            for uscd_match in self._cf.options("update_sheet:条件:匹配"):
                d_match[uscd_match] = self._cf.get("update_sheet:条件:匹配", uscd_match).strip()
            self._dict_update_sheet_condition['match'] = d_match

            d_break = {}
            for uscd_break in self._cf.options("update_sheet:条件:停止"):
                d_break[uscd_break] = self._cf.get("update_sheet:条件:停止", uscd_break).strip()
            self._dict_update_sheet_condition['break'] = d_break

            logger.debug("add_sheet:%s,add_sheet_column:%s,update_sheet:%s,update_sheet_column:%s,update_sheet_conditio:%s",
                         self._dict_title_add_sheet,self._dict_add_sheet_column,
                         self._dict_title_update_sheet,self._dict_update_sheet_column, self._dict_update_sheet_condition)

        except BaseException as e:
            logger.exception(e)
            raise e

    #@profile
    def make_the_same_style(self, sheet):
        logger.debug("表名：%s,min_row:%s,max_row:%s,min_column:%s,max_column:%s",
                     sheet.title, sheet.min_row, sheet.max_row, sheet.min_column, sheet.max_column)

        if sheet.max_row < 4:
            return

        for i in range(1, min(sheet.max_column + 1, 100)):
            sheet.cell(row=sheet.max_row, column=i).font = sheet.cell(row=sheet.max_row - 1, column=i).font.copy()
            sheet.cell(row=sheet.max_row, column=i).fill = sheet.cell(row=sheet.max_row - 1, column=i).fill.copy()
            sheet.cell(row=sheet.max_row, column=i).border = sheet.cell(row=sheet.max_row - 1, column=i).border.copy()
            sheet.cell(row=sheet.max_row, column=i).alignment = sheet.cell(row=sheet.max_row - 1,
                                                                           column=i).alignment.copy()
            sheet.cell(row=sheet.max_row, column=i).number_format = sheet.cell(row=sheet.max_row - 1,
                                                                               column=i).number_format

    #@profile
    def get_title_values(self, line):
        # 每一行有且只有一个标题：审核结果/还款/续期
        titles = re.findall(r'{0}'.format(self._cf.get("re", "title").strip()), line)
        if len(titles) != 1:
            logger.error("记录内容有误，没能找到标题，记录内容为：%s")
            return None, None

        if not titles[0] in self._dict_title_values_count:
            logger.error("没能在配置文件：%s 找到对应于标题为：%s 的个数配置",
                         auto_excel_config, titles[0])
            return None, None

        # 提取记录中的数据并做检查
        values = re.findall(r'{0}'.format(self._cf.get("re", "values").strip()), line)
        if len(values) != self._dict_title_values_count[titles[0]]:
            logger.error("在记录中提取到的数据个数：%s 与配置的个数：%s 不一致，判定记录内容有误：%s",
                         len(values), self._dict_title_values_count[titles[0]], line)
            return None, None

        #数值类型修正
        if titles[0] in self._dict_title_float_pos:
            for t in self._dict_title_float_pos[titles[0]]:
                values[t] = float(values[t])

        if titles[0] in self._dict_check and not eval(self._dict_check[titles[0]]):
            logger.error("记录无法通过配置中对标题：%s的校验，记录内容为：%s", titles[0], line)
            return None, None

        return titles[0], values

    #@profile
    def add_to_sheet(self, book, title, values):
        logger.info("标题：%s, 提取的数据：%s", title, values)

        if title not in self._dict_title_add_sheet:
            logger.warning("无法在配置中找到对应标题：%s的excel表名配置，将直接使用该标题作为操作的excel表名", title)
            sheet_name = title
        else:
            sheet_name = self._dict_title_add_sheet[title]

        if sheet_name not in book.get_sheet_names():
            book.create_sheet(sheet_name)
        sheet = book.get_sheet_by_name(sheet_name)
        logger.debug("表名：%s, 维度：%s", sheet_name, sheet.dimensions)

        if title not in self._dict_add_sheet_column:
            logger.warn("无法在配置中找到对应标题：%s的excel列关系配置，将直接使用原始数据内容和顺序进行表格填充", title)
            values_to_add = values
        else:
            values_to_add = ["" for x in range(max(self._dict_add_sheet_column[title].keys()) + 1)]

        for k,v in self._dict_add_sheet_column[title].items():
            if v.isdigit():
                values_to_add[k] = values[int(v)]
            else:
                values_to_add[k] = v.format(sheet.max_row + 1)

        sheet.append(values_to_add)
        logger.debug("写入excel表：%s，数据：%s", title, values_to_add)

        self.make_the_same_style(sheet)

    #@profile
    def find_row_to_update(self, sheet, title, values):
        if title not in self._dict_update_sheet_condition['match']:
            logger.error("无法在配置中找到对应标题：%s的更新记录需满足的条件", title)
            return -1

        row_to_update = -1
        # 从下往上找到满足条件的行
        for i in range(sheet.max_row, 0, -1):
            if eval(self._dict_update_sheet_condition['match'][title]):
                row_to_update = i
                break

        if row_to_update < 0 or title not in self._dict_update_sheet_condition['break']:
            return row_to_update

        # 有可能存在连发2个审核，后带一个还款的情况，再往前搜索直到遇到停止条件
        for i in range(row_to_update - 1, 0, -1):
            if eval(self._dict_update_sheet_condition['match'][title]):
                row_to_update = i
                break
            elif eval(self._dict_update_sheet_condition['break'][title]):
                break
        return row_to_update

    #@profile
    def update_sheet(self, book, title, values):
        if title not in self._dict_title_update_sheet:
            logger.debug("标题：%s不在需要更新excel表名的配置中，不需要更新到excel中", title)
            return

        sheet = book.get_sheet_by_name(self._dict_title_update_sheet[title])
        row_to_update = self.find_row_to_update(sheet, title, values)

        if row_to_update < 1:
            logger.error("无法找到满足条件需要更新的行，标题：%s,数据：%s", title, values)
            return

        logger.debug("row_to_update:%s", row_to_update)

        if title not in self._dict_update_sheet_column:
            logger.error("无法在配置中找到对应标题：%s需要更新的列", title)
            return

        for k,v in self._dict_update_sheet_column[title].items():
            cl = sheet.cell(row = row_to_update, column = k + 1)
            l = v.split(':')
            op = l[0].strip()
            index = int(l[1].strip())
            if op == '=':
                cl.value = values[index]
            elif op == '+':
                if cl.value is None or cl.value == '':
                    cl.value = values[index]
                else:
                    cl.value += values[index]

    #@profile
    def do(self):
        start = time.time()
        book = load_workbook(self._xlsx_file_path)
        txt_lines = open(self._txt_file_path, mode='r', encoding='UTF-8')

        line_count = 0
        for line in txt_lines:
            title, values = self.get_title_values(line)
            logger.debug("title:%s, values:%s", title, values)

            if title is None or values is None:
                logger.error("记录有误，中断本次录入")
                return wait_to_quit()

            self.add_to_sheet(book, title, values)
            self.update_sheet(book, title, values)
            line_count += 1

        book.save(self._xlsx_file_path)
        logger.info("处理%s条记录，耗时%s秒", line_count, round(time.time() - start, 3))
        return wait_to_quit()


if __name__ == '__main__':
    baseconfdir = "config"
    loggingconf = "log.config"
    auto_excel_config = "auto.ini"

    try:
        logging.config.fileConfig(os.path.join(os.getcwd(), baseconfdir, loggingconf))
        logger = logging.getLogger()

        auto_excel = AutoExcel(auto_excel_config)
        auto_excel.do()
    except BaseException as e:
        logger.exception(e)
        wait_to_quit()
