# -*- coding: utf-8 -*-
__author__ = 'xujh'

#打包 pyinstaller -F --upx-dir D:\soft\upx391w -i excel_72pxt.ico auto_excel.py
import os
import logging.config
import configparser
from openpyxl import *
import re
from openpyxl.styles import *
#import copy
import time
#from .check import *

logger = logging.getLogger()

# def wait_to_quit():
#     print("按回车键结束")
#     if input():
#         exit()


class Excel(object):
    def __init__(self, cf, check):
        try:
            self._cf = cf
            self._check = check
            self._txt_file_path = self._cf.get("path", "txt_file_path")
            if not os.path.exists(self._txt_file_path):
                logger.error("txt文件不存在，文件路径配置为:%s", self._txt_file_path)
                raise IOError

            self._xlsx_file_path = self._cf.get("path", "xlsx_file_path")
            if not os.path.exists(self._xlsx_file_path):
                logger.info("xlsx文件不存在，将创建一个新的，文件路径配置为：%s", self._xlsx_file_path)
                dirname = os.path.dirname(self._xlsx_file_path)
                if not os.path.exists(dirname):
                    os.mkdir(dirname)
                Workbook().save(self._xlsx_file_path)

            self._book = load_workbook(self._xlsx_file_path)

            # 记录类型及需要追加记录的excel表格名
            self._dict_title_add_sheet = {}
            # 记录类型及需要写入的列数据
            self._dict_add_sheet_column = {}
            # 记录类型及需要更新记录的excel表格名
            self._dict_title_update_sheet = {}
            # 记录类型及需要更新的列
            self._dict_update_sheet_column = {}
            # 更新的记录需要满足的条件
            self._dict_update_sheet_condition = {}


            for asn in self._cf.options("add_sheet"):
                self._dict_title_add_sheet[asn] = self._cf.get("add_sheet", asn).strip()
                d = {}
                for asc in self._cf.options("add_sheet:" + asn):
                    d[int(asc)] = self._cf.get("add_sheet:" + asn, asc).strip()
                self._dict_add_sheet_column[asn] = d

            for usn in self._cf.options("update_sheet"):
                self._dict_title_update_sheet[usn] = self._cf.get("update_sheet", usn).strip()
                d = {}
                for usc in self._cf.options("update_sheet:" + usn):
                    d[int(usc)] = self._cf.get("update_sheet:" + usn, usc).strip()
                self._dict_update_sheet_column[usn] = d

            d_match = {}
            for uscd_match in self._cf.options("update_sheet:条件:匹配"):
                d_match[uscd_match] = self._cf.get("update_sheet:条件:匹配", uscd_match).strip()
            self._dict_update_sheet_condition['match'] = d_match

            d_break = {}
            for uscd_break in self._cf.options("update_sheet:条件:停止"):
                d_break[uscd_break] = self._cf.get("update_sheet:条件:停止", uscd_break).strip()
            self._dict_update_sheet_condition['break'] = d_break

            logger.debug("add_sheet:%s,add_sheet_column:%s,update_sheet:%s,update_sheet_column:%s,update_sheet_conditio:%s",
                         self._dict_title_add_sheet,self._dict_add_sheet_column,
                         self._dict_title_update_sheet,self._dict_update_sheet_column, self._dict_update_sheet_condition)

        except BaseException as e:
            logger.exception(e)
            raise e

    #@profile
    def make_the_same_style(self, sheet):
        logger.debug("表名：%s,min_row:%s,max_row:%s,min_column:%s,max_column:%s",
                     sheet.title, sheet.min_row, sheet.max_row, sheet.min_column, sheet.max_column)

        if sheet.max_row < 4:
            return

        for i in range(1, min(sheet.max_column + 1, 100)):
            sheet.cell(row=sheet.max_row, column=i).font = sheet.cell(row=sheet.max_row - 1, column=i).font.copy()
            sheet.cell(row=sheet.max_row, column=i).fill = sheet.cell(row=sheet.max_row - 1, column=i).fill.copy()
            sheet.cell(row=sheet.max_row, column=i).border = sheet.cell(row=sheet.max_row - 1, column=i).border.copy()
            sheet.cell(row=sheet.max_row, column=i).alignment = sheet.cell(row=sheet.max_row - 1,
                                                                           column=i).alignment.copy()
            sheet.cell(row=sheet.max_row, column=i).number_format = sheet.cell(row=sheet.max_row - 1,
                                                                               column=i).number_format

    #@profile
    def add_to_sheet(self, title, values):
        logger.info("标题：%s, 提取的数据：%s", title, values)

        if title not in self._dict_title_add_sheet:
            logger.warning("无法在配置中找到对应标题：%s的excel表名配置，将直接使用该标题作为操作的excel表名", title)
            sheet_name = title
        else:
            sheet_name = self._dict_title_add_sheet[title]

        if sheet_name not in self._book.sheetnames:
            self._book.create_sheet(sheet_name)
        sheet = self._book[sheet_name]
        logger.debug("表名：%s, 维度：%s", sheet_name, sheet.dimensions)

        if title not in self._dict_add_sheet_column:
            logger.warn("无法在配置中找到对应标题：%s的excel列关系配置，将直接使用原始数据内容和顺序进行表格填充", title)
            values_to_add = values
        else:
            values_to_add = ["" for x in range(max(self._dict_add_sheet_column[title].keys()) + 1)]

        for k,v in self._dict_add_sheet_column[title].items():
            if v.isdigit():
                values_to_add[k] = values[int(v)]
            else:
                values_to_add[k] = v.format(sheet.max_row + 1)

        sheet.append(values_to_add)
        logger.debug("写入excel表：%s，数据：%s", title, values_to_add)

        self.make_the_same_style(sheet)

    #@profile
    def find_row_to_update(self, sheet, title, values):
        if title not in self._dict_update_sheet_condition['match']:
            #logger.error("无法在配置中找到对应标题：%s的更新记录需满足的条件", title)
            return -1

        row_to_update = -1
        # 从下往上找到满足条件的行
        for i in range(sheet.max_row, 0, -1):
            if eval(self._dict_update_sheet_condition['match'][title]):
                row_to_update = i
                break

        if row_to_update < 0 or title not in self._dict_update_sheet_condition['break']:
            return row_to_update

        # 有可能存在连发2个审核，后带一个还款的情况，再往前搜索直到遇到停止条件
        for i in range(row_to_update - 1, 0, -1):
            if eval(self._dict_update_sheet_condition['match'][title]):
                row_to_update = i
                break
            elif eval(self._dict_update_sheet_condition['break'][title]):
                break
        return row_to_update

    #@profile
    def update_sheet(self, title, values):
        if title not in self._dict_title_update_sheet:
            logger.debug("标题为：%s 的记录不需要做更新操作", title)
            return True

        sheet = self._book.get_sheet_by_name(self._dict_title_update_sheet[title])
        row_to_update = self.find_row_to_update(sheet, title, values)

        if row_to_update < 1:
            logger.error("无法找到满足条件需要更新的行，标题：%s,数据：%s", title, values)
            return False

        logger.debug("row_to_update:%s", row_to_update)

        if title not in self._dict_update_sheet_column:
            logger.error("无法在配置中找到对应标题：%s需要更新的列", title)
            return False

        for k,v in self._dict_update_sheet_column[title].items():
            cl = sheet.cell(row = row_to_update, column = k + 1)
            l = v.split(':')
            op = l[0].strip()
            index = int(l[1].strip())
            if op == '=':
                cl.value = values[index]
            elif op == '+':
                if cl.value is None or cl.value == '':
                    cl.value = values[index]
                else:
                    cl.value += values[index]
        return True

    def upsert_by_line(self, line):
        title, values, rtn_str = self._check.get_title_values(line)
        logger.debug("title:%s, values:%s, rtn_str:%s", title, values, rtn_str)

        if title is None or values is None:
            logger.error("记录有误，中断本次录入")
            return False

        self.add_to_sheet(title, values)
        if not self.update_sheet(title, values):
            return False
        return True

    def upsert_by_values(self, title, values):
        self.add_to_sheet(title, values)
        self.update_sheet(title, values)

    #@profile
    def upsert_from_txt(self):
        start = time.time()
        logger.info("将读取txt文件：%s，并导入到excel文件中：%s", self._txt_file_path, self._xlsx_file_path)
        txt_lines = open(self._txt_file_path, mode='r', encoding='UTF-8')

        line_count = 0
        for line in txt_lines:
            if not self.upsert_by_line(line):
                return
            line_count += 1

        self._book.save(self._xlsx_file_path)
        logger.info("处理%s条记录，耗时%s秒", line_count, round(time.time() - start, 3))
        #return wait_to_quit()

    def save(self):
        self._book.save(self._xlsx_file_path)
    # def upsert_from_line(self, line):
    #     if self.upsert_by_line(line):
    #         self._book.save(self._xlsx_file_path)


# if __name__ == '__main__':
#     baseconfdir = "./config"
#     loggingconf = "log.config"
#     config = "auto.ini"
#
#     try:
#         print(os.getcwd())
#         logging.config.fileConfig(os.path.join(os.getcwd(), baseconfdir, loggingconf))
#         logger = logging.getLogger()
#
#         cf = configparser.ConfigParser()
#         cf.read(os.path.join(os.getcwd(), baseconfdir, config), encoding='UTF-8')
#
#         auto_excel = Excel(cf)
#         auto_excel.upsert_from_txt()
#     except BaseException as e:
#         logger.exception(e)
#         wait_to_quit()
