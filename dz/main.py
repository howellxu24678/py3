# -*- coding: utf-8 -*-
__author__ = 'xujh'

# s = "审核结果：借款人[马应娇]。合同金额[1200]元，到账金额[1000]元，期限[10]天，借款日期[2018.01.16]，" \
#     "还款日期[2018.01.26]，到期还款金额[1200]元，逾期管理费[30]元/天，随时可以提前还款，提前还款不收取违约金。"
#
#
#
# s3 = "借款人[]。合同金额[]元，到账金额[]元，期限[]天，借款日期[]，还款日期[]，到期还款金额[]元，逾期管理费[]元/天，随时可以提前还款，提前还款不收取违约金。"
#
# import re
#
# l = re.findall(r'\[(.*?)\]', s)
#
# s1 = "【审核结果】：借款人【马应娇】。合同金额【1200】元"
# l1 = re.findall(r'\【(.*?)\】', s1)

import os
import logging.config
import configparser
import openpyxl
import re

baseconfdir = "config"
loggingconf = "log.config"
businessconf = "dz.ini"

logging.config.fileConfig(os.path.join(os.getcwd(), baseconfdir, loggingconf))
logger = logging.getLogger()

cf = configparser.ConfigParser()
cf.read(os.path.join(os.getcwd(), baseconfdir, businessconf), encoding='UTF-8')


try:
    txt_file_path = cf.get("path", "txt_file_path")
    if not os.path.exists(txt_file_path):
        logger.error("txt文件不存在，文件路径配置为:%s", txt_file_path)
        exit()

    xlsx_file_path = cf.get("path", "xlsx_file_path")
    if not os.path.exists(xlsx_file_path):
        logger.info("xlsx文件不存在，将创建一个新的，文件路径配置为：%s", xlsx_file_path)
        dirname = os.path.dirname(xlsx_file_path)
        if not os.path.exists(dirname):
            os.mkdir(dirname)
        openpyxl.Workbook().save(xlsx_file_path)

    #每个类型的记录对应提取的结果集个数（用于检查记录是否正确）
    d_count = {}
    for c in cf.options("count"):
        d_count[c] = cf.getint("count", c)

    #结果集为数值的位置
    d_type = {}
    for t in cf.options("type"):
        d_type[t] = [int(x) for x in cf.get("type", t).split(',')]

    wb = openpyxl.load_workbook(xlsx_file_path)
    fpy = open(txt_file_path, mode='r', encoding='UTF-8')
    for line in fpy:
        #每一行有且只有一个标题：审核结果/还款/续期
        title = re.findall(r'\【(.*?)\】', line)
        if len(title) != 1:
            logger.error("记录内容有误，没能找到标题，记录内容为：%s")
            continue

        if not title[0] in d_count:
            logger.error("没能在配置文件：%s 找到对应于标题为：%s 的个数配置",
                         businessconf, title[0])
            continue

        #提取记录中的数据并做检查
        values = re.findall(r'\[(.*?)\]', line)
        if len(values) != d_count[title[0]]:
            logger.error("在记录中提取到的数据个数：%s 与配置的个数：%s 不一致，判定记录内容有误：%s",
                         len(values), d_count[title[0]], line)
            continue

        if title[0] in d_type:
            for t in d_type[title[0]]:
                values[t] = float(values[t])

        logger.info("记录标题：%s, 记录数据：%s",title[0], values)

        if not title[0] in wb.get_sheet_names():
            wb.create_sheet(title[0])
        ws = wb.get_sheet_by_name(title[0])
        ws.append(values)

    wb.save(xlsx_file_path)

except BaseException as e:
    logger.exception(e)