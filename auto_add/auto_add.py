# -*- coding: utf-8 -*-
__author__ = 'xujh'

import os
import logging.config
import configparser
from openpyxl import *
import re
from openpyxl.styles import *
import copy

baseconfdir = "config"
loggingconf = "log.config"
businessconf = "auto_add.ini"

logging.config.fileConfig(os.path.join(os.getcwd(), baseconfdir, loggingconf))
logger = logging.getLogger()

cf = configparser.ConfigParser()
cf.read(os.path.join(os.getcwd(), baseconfdir, businessconf), encoding='UTF-8')


#前后两行的的样式一致
def make_thesame_style(sheet):
    logger.debug("表名：%s,min_row:%s,max_row:%s,min_column:%s,max_column:%s",
                 sheet.title, sheet.min_row, sheet.max_row, sheet.min_column, sheet.max_column)

    if sheet.max_row < 4:
        return

    for i in range(1, min(sheet.max_column + 1, 100)):
        sheet.cell(row=sheet.max_row, column=i).font = sheet.cell(row=sheet.max_row - 1, column=i).font.copy()
        sheet.cell(row=sheet.max_row, column=i).fill = sheet.cell(row=sheet.max_row - 1, column=i).fill.copy()
        sheet.cell(row=sheet.max_row, column=i).border = sheet.cell(row=sheet.max_row - 1, column=i).border.copy()
        sheet.cell(row=sheet.max_row, column=i).alignment = sheet.cell(row=sheet.max_row - 1, column=i).alignment.copy()
        sheet.cell(row=sheet.max_row, column=i).number_format = sheet.cell(row=sheet.max_row - 1, column=i).number_format

#解析记录中的标题和数据，并做检查
def get_title_values(line, d_count, d_type):
    # 每一行有且只有一个标题：审核结果/还款/续期
    titles = re.findall(r'\【(.*?)\】', line)
    if len(titles) != 1:
        logger.error("记录内容有误，没能找到标题，记录内容为：%s")
        return None,None

    if not titles[0] in d_count:
        logger.error("没能在配置文件：%s 找到对应于标题为：%s 的个数配置",
                     businessconf, titles[0])
        return None, None

    # 提取记录中的数据并做检查
    values = re.findall(r'\[(.*?)\]', line)
    if len(values) != d_count[titles[0]]:
        logger.error("在记录中提取到的数据个数：%s 与配置的个数：%s 不一致，判定记录内容有误：%s",
                     len(values), d_count[titles[0]], line)
        return None, None

    if titles[0] in d_type:
        for t in d_type[titles[0]]:
            values[t] = float(values[t])

    return titles[0], values


#记录写入表格
def write_to_sheet(title, values):
    logger.info("标题：%s, 提取的数据：%s", title, values)

    if not title in book.get_sheet_names():
        book.create_sheet(title)
    sheet = book.get_sheet_by_name(title)
    logger.debug("表名：%s, 维度：%s", title, sheet.dimensions)

    # 前面加一列空值
    values_ = copy.copy(values)
    values_.insert(0, '')
    if title == '审核结果':
        # 次数
        values_.insert(2, '=IF(B{0}="","",COUNTIF($A$3:B{0},TEXT(B{0},"###")))'.format(sheet.max_row + 1))
        # 手机号码
        values_.insert(3,
                      '=IFERROR(IF(VLOOKUP(B{0},客户信息表!B:D,2,FALSE)="","",VLOOKUP(B{0},客户信息表!B:D,2,FALSE)),"")'.format(
                          sheet.max_row + 1))
        # 所在微信
        values_.insert(4,
                      '=IFERROR(IF(VLOOKUP(B{0},客户信息表!B:D,3,FALSE)="","",VLOOKUP(B{0},客户信息表!B:D,3,FALSE)),"")'.format(
                          sheet.max_row + 1))
        # 应还总额
        values_.insert(7, '=IF(N{0}="已结清",0,F{0})'.format(sheet.max_row + 1))
        values_ = values_[0:11]

    sheet.append(values_)
    logger.debug("写入excel表：%s，数据：%s", title, values_)

    make_thesame_style(sheet)


def modify_data(book, title, values):
    if title == '审核结果':
        return

    sheet = book.get_sheet_by_name('审核结果')
    for i in range(sheet.max_row, 1, -1):
        #从下往上找到姓名相同的那一行
        if sheet.cell(row=i, column=2).value == values[1]:
            #已还金额增加（还款和续期一致处理）
            if sheet.cell(row=i, column=13).value is None:
                sheet.cell(row=i, column=13).value = values[2]
            else:
                sheet.cell(row=i, column=13).value += values[2]

            if title == '还款':
                sheet.cell(row=i, column=12).value = values[0]
            elif title == '续期':
                sheet.cell(row=i, column=12).value = values[3]

            break



try:
    txt_file_path = cf.get("path", "txt_file_path")
    if not os.path.exists(txt_file_path):
        logger.error("txt文件不存在，文件路径配置为:%s", txt_file_path)
        exit()

    xlsx_file_path = cf.get("path", "xlsx_file_path")
    if not os.path.exists(xlsx_file_path):
        logger.info("xlsx文件不存在，将创建一个新的，文件路径配置为：%s", xlsx_file_path)
        dirname = os.path.dirname(xlsx_file_path)
        if not os.path.exists(dirname):
            os.mkdir(dirname)
        Workbook().save(xlsx_file_path)

    #每个类型的记录对应提取的结果集个数（用于检查记录是否正确）
    d_count = {}
    for c in cf.options("count"):
        d_count[c] = cf.getint("count", c)

    #结果集为数值的位置
    d_type = {}
    for t in cf.options("type"):
        d_type[t] = [int(x) for x in cf.get("type", t).split(',')]

    book = load_workbook(xlsx_file_path)
    fpy = open(txt_file_path, mode='r', encoding='UTF-8')
    for line in fpy:
        title, values = get_title_values(line, d_count, d_type)
        write_to_sheet(title, values)
        modify_data(book, title, values)

    book.save(xlsx_file_path)

except BaseException as e:
    logger.exception(e)